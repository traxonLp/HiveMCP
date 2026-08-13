"""Render a SheetSpec into a .xlsx file."""

from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.chart import AreaChart, BarChart, LineChart, PieChart, Reference, ScatterChart
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, Rule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

from .base import MEDIA_TYPES, RenderedFile, RenderError
from .theme import RenderWarnings, check_font, normalize_hex, safe_filename

NUMBER_FORMATS = {
    "text": "@",
    "number": "#,##0.00",
    "integer": "#,##0",
    "currency": '#,##0.00 "€"',
    "percent": "0.0%",
    "date": "DD.MM.YYYY",
    "bool": "General",
    "formula": "General",
}

# Excel treats a leading one of these in a *text* cell as the start of a formula.
# A spreadsheet built from model output or an uploaded document is untrusted input,
# so anything not explicitly typed as a formula gets neutralised.
FORMULA_TRIGGERS = ("=", "+", "-", "@", "\t", "\r")

DEFAULT_HEADER_FILL = "DDE5F0"
MAX_COLUMN_WIDTH = 60.0
MIN_COLUMN_WIDTH = 8.0


class XlsxRenderer:
    def __init__(self, options: Any, template_path: Path | None = None) -> None:
        self.options = options
        self.template_path = template_path
        self.warnings = RenderWarnings()
        self.font = check_font(
            getattr(options, "font_family", None),
            self.warnings,
            getattr(options, "language", "en"),
        )

    # ---------------------------------------------------------------- public

    def render(self, spec: Any) -> RenderedFile:
        workbook = self._open_workbook()
        existing = set(workbook.sheetnames)

        for index, sheet_spec in enumerate(spec.sheets):
            try:
                worksheet = workbook.create_sheet(title=sheet_spec.name)
                self._write_sheet(worksheet, sheet_spec)
            except Exception as exc:  # noqa: BLE001 - a bad sheet must name itself
                raise RenderError(
                    f"sheet {index + 1} ({sheet_spec.name!r}) could not be rendered: {exc}"
                ) from exc

        # Drop whatever the blank workbook or template started with, but only once at
        # least one real sheet exists: a workbook with zero sheets is invalid.
        for name in existing:
            if len(workbook.sheetnames) > 1:
                del workbook[name]

        workbook.properties.title = spec.title

        buffer = BytesIO()
        workbook.save(buffer)
        return RenderedFile(
            data=buffer.getvalue(),
            filename=safe_filename(getattr(self.options, "filename", None), spec.title, "xlsx"),
            media_type=MEDIA_TYPES["xlsx"],
            warnings=list(self.warnings),
            sheet_names=list(workbook.sheetnames),
        )

    # --------------------------------------------------------------- private

    def _open_workbook(self) -> Workbook:
        if self.template_path is None:
            return Workbook()
        try:
            return load_workbook(str(self.template_path))
        except Exception as exc:  # noqa: BLE001
            raise RenderError(
                f"template {self.template_path.name!r} could not be opened as an Excel "
                f"file: {exc}"
            ) from exc

    def _write_sheet(self, worksheet: Worksheet, spec: Any) -> None:
        self._write_header(worksheet, spec.columns)
        self._write_rows(worksheet, spec.columns, spec.rows)
        self._size_columns(worksheet, spec.columns, spec.rows)

        last_column = get_column_letter(len(spec.columns))
        last_row = len(spec.rows) + 1

        if spec.freeze_header:
            worksheet.freeze_panes = "A2"
        if spec.autofilter and spec.rows:
            worksheet.auto_filter.ref = f"A1:{last_column}{last_row}"

        for rule_spec in spec.conditional_formats:
            self._add_conditional_format(worksheet, rule_spec)
        for chart_spec in spec.charts:
            self._add_chart(worksheet, chart_spec)

    def _write_header(self, worksheet: Worksheet, columns: list[Any]) -> None:
        fill_hex = normalize_hex(
            (getattr(self.options, "theme_colors", None) or {}).get("accent1")
        ) or DEFAULT_HEADER_FILL
        fill = PatternFill("solid", fgColor=fill_hex)
        font = Font(bold=True, name=self.font or None)
        border = Border(bottom=Side(style="thin", color="FF9E9E9E"))

        for index, column in enumerate(columns, start=1):
            cell = worksheet.cell(row=1, column=index, value=column.header)
            cell.fill = fill
            cell.font = font
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        worksheet.row_dimensions[1].height = 20

    def _write_rows(self, worksheet: Worksheet, columns: list[Any], rows: list[dict]) -> None:
        body_font = Font(name=self.font) if self.font else None
        for row_index, row in enumerate(rows, start=2):
            for column_index, column in enumerate(columns, start=1):
                raw = row.get(column.key)
                cell = worksheet.cell(row=row_index, column=column_index)
                cell.value = self._coerce(raw, column.type)
                cell.number_format = column.number_format or NUMBER_FORMATS[column.type]
                if body_font is not None:
                    cell.font = body_font

    @staticmethod
    def _coerce(value: Any, cell_type: str) -> Any:
        if value is None:
            return None
        if cell_type == "formula":
            text = str(value)
            return text if text.startswith("=") else f"={text}"
        if cell_type in {"number", "integer", "percent", "currency"}:
            try:
                number = float(value)
            except (TypeError, ValueError):
                return XlsxRenderer._neutralise(str(value))
            return int(number) if cell_type == "integer" and number.is_integer() else number
        if cell_type == "bool":
            if isinstance(value, bool):
                return value
            return str(value).strip().lower() in {"true", "wahr", "ja", "yes", "1"}
        if isinstance(value, str):
            return XlsxRenderer._neutralise(value)
        return value

    @staticmethod
    def _neutralise(text: str) -> str:
        """Stop Excel from evaluating untrusted text as a formula."""
        return f"'{text}" if text[:1] in FORMULA_TRIGGERS else text

    def _size_columns(self, worksheet: Worksheet, columns: list[Any], rows: list[dict]) -> None:
        for index, column in enumerate(columns, start=1):
            letter = get_column_letter(index)
            if column.width:
                worksheet.column_dimensions[letter].width = column.width
                continue
            widest = len(column.header)
            for row in rows[:500]:  # sampling is enough, and keeps big sheets fast
                value = row.get(column.key)
                if value is not None:
                    widest = max(widest, len(str(value)))
            worksheet.column_dimensions[letter].width = min(
                MAX_COLUMN_WIDTH, max(MIN_COLUMN_WIDTH, widest + 2)
            )

    def _add_conditional_format(self, worksheet: Worksheet, spec: Any) -> None:
        highlight = DifferentialStyle(fill=PatternFill(bgColor="FFC7CE"), font=Font(color="9C0006"))
        rules = {
            "color_scale": lambda: ColorScaleRule(
                start_type="min",
                start_color="F8696B",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFEB84",
                end_type="max",
                end_color="63BE7B",
            ),
            "data_bar": lambda: DataBarRule(
                start_type="min", end_type="max", color="638EC6", showValue=True
            ),
            "above_average": lambda: Rule(type="aboveAverage", dxf=highlight),
            "below_average": lambda: Rule(type="aboveAverage", aboveAverage=False, dxf=highlight),
            "duplicates": lambda: Rule(type="duplicateValues", dxf=highlight),
        }
        factory = rules.get(spec.rule)
        if factory is None:
            self.warnings.add(f"Unknown conditional format rule {spec.rule!r}; skipped.")
            return
        worksheet.conditional_formatting.add(spec.range, factory())

    def _add_chart(self, worksheet: Worksheet, spec: Any) -> None:
        builders = {
            "bar": lambda: self._bar_chart("bar"),
            "column": lambda: self._bar_chart("col"),
            "line": LineChart,
            "pie": PieChart,
            "scatter": ScatterChart,
            "area": AreaChart,
        }
        chart = builders[spec.chart_type]()
        if spec.title:
            chart.title = spec.title

        try:
            values = self._reference(worksheet, spec.values_range)
            categories = self._reference(worksheet, spec.categories_range)
        except ValueError as exc:
            raise RenderError(f"chart range is not valid A1 notation: {exc}") from exc

        chart.add_data(values, titles_from_data=True)
        chart.set_categories(categories)
        chart.height = 8
        chart.width = 16
        worksheet.add_chart(chart, spec.anchor)

    @staticmethod
    def _bar_chart(direction: str) -> BarChart:
        chart = BarChart()
        chart.type = direction
        return chart

    @staticmethod
    def _reference(worksheet: Worksheet, a1_range: str) -> Reference:
        min_col, min_row, max_col, max_row = range_boundaries(a1_range)
        return Reference(
            worksheet, min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row
        )


def render_spreadsheet(
    spec: Any, options: Any, template_path: Path | None = None
) -> RenderedFile:
    return XlsxRenderer(options, template_path).render(spec)
