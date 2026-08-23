"""Excel rendering."""

from __future__ import annotations

from io import BytesIO

import pytest
from openpyxl import load_workbook

from hivemcp.core.models import (
    Column,
    ConditionalFormat,
    RenderOptions,
    Sheet,
    SheetChart,
    SheetSpec,
)
from hivemcp.core.render.base import MEDIA_TYPES
from hivemcp.core.render.xlsx import NUMBER_FORMATS, render_spreadsheet


def opened(rendered):
    return load_workbook(BytesIO(rendered.data))


# --------------------------------------------------------------------------- #
# Structure
# --------------------------------------------------------------------------- #


def test_renders_sheets_headers_and_rows(
    workbook: SheetSpec, options: RenderOptions
) -> None:
    book = opened(render_spreadsheet(workbook, options))
    sheet = book["Regionen"]
    assert [cell.value for cell in sheet[1]] == ["Region", "Umsatz"]
    assert sheet["A2"].value == "DACH"
    assert sheet["B2"].value == 4200000


def test_reports_sheet_names_and_media_type(
    workbook: SheetSpec, options: RenderOptions
) -> None:
    rendered = render_spreadsheet(workbook, options)
    assert rendered.sheet_names == ["Regionen"]
    assert rendered.media_type == MEDIA_TYPES["xlsx"]
    assert rendered.filename.endswith(".xlsx")


def test_output_is_a_real_ooxml_package(
    workbook: SheetSpec, options: RenderOptions
) -> None:
    assert render_spreadsheet(workbook, options).data[:2] == b"PK"


def test_several_sheets_keep_their_order(options: RenderOptions) -> None:
    spec = SheetSpec(
        title="T",
        sheets=[
            Sheet(name=name, columns=[Column(header="A", key="a")], rows=[{"a": "x"}])
            for name in ("Erste", "Zweite", "Dritte")
        ],
    )
    assert opened(render_spreadsheet(spec, options)).sheetnames == [
        "Erste",
        "Zweite",
        "Dritte",
    ]


def test_missing_keys_leave_cells_empty_rather_than_shifting_columns(
    options: RenderOptions,
) -> None:
    """A row that omits a key must not slide the remaining values left."""
    spec = SheetSpec(
        title="T",
        sheets=[
            Sheet(
                name="S",
                columns=[
                    Column(header="A", key="a"),
                    Column(header="B", key="b"),
                    Column(header="C", key="c"),
                ],
                rows=[{"a": "1", "c": "3"}],
            )
        ],
    )
    sheet = opened(render_spreadsheet(spec, options))["S"]
    assert sheet["A2"].value == "1"
    assert sheet["B2"].value in (None, "")
    assert sheet["C2"].value == "3"


# --------------------------------------------------------------------------- #
# Types and formats
# --------------------------------------------------------------------------- #


@pytest.mark.parametrize(
    ("cell_type", "value", "expected"),
    [
        ("integer", 42, 42),
        ("number", 1.5, 1.5),
        ("currency", 1000, 1000),
        ("percent", 0.25, 0.25),
        ("bool", True, True),
        ("text", "hallo", "hallo"),
    ],
)
def test_cell_types_keep_their_native_value(
    cell_type: str, value, expected, options: RenderOptions
) -> None:
    spec = SheetSpec(
        title="T",
        sheets=[
            Sheet(
                name="S",
                columns=[Column(header="H", key="k", type=cell_type)],
                rows=[{"k": value}],
            )
        ],
    )
    sheet = opened(render_spreadsheet(spec, options))["S"]
    assert sheet["A2"].value == expected
    assert sheet["A2"].number_format == NUMBER_FORMATS[cell_type]


def test_formula_cells_are_written_as_formulas(options: RenderOptions) -> None:
    spec = SheetSpec(
        title="T",
        sheets=[
            Sheet(
                name="S",
                columns=[Column(header="Summe", key="s", type="formula")],
                rows=[{"s": "=1+1"}],
            )
        ],
    )
    sheet = opened(render_spreadsheet(spec, options))["S"]
    assert sheet["A2"].value == "=1+1"


@pytest.mark.parametrize("payload", ["=1+1", "+1", "-1", "@SUM(A1)"])
def test_text_cells_cannot_smuggle_in_a_formula(
    payload: str, options: RenderOptions
) -> None:
    """Untrusted text starting with a trigger character must not be evaluated by Excel.

    This is the injection path: content arrives from a model or an uploaded document, and
    a plain-text cell beginning with ``=`` would otherwise run as a formula on open.
    """
    spec = SheetSpec(
        title="T",
        sheets=[
            Sheet(
                name="S",
                columns=[Column(header="H", key="k", type="text")],
                rows=[{"k": payload}],
            )
        ],
    )
    value = opened(render_spreadsheet(spec, options))["S"]["A2"].value
    assert value.startswith("'") or not value.startswith(("=", "+", "-", "@"))


# --------------------------------------------------------------------------- #
# Sheet features
# --------------------------------------------------------------------------- #


def test_header_row_is_frozen_when_requested(options: RenderOptions) -> None:
    spec = SheetSpec(
        title="T",
        sheets=[
            Sheet(
                name="S",
                columns=[Column(header="A", key="a")],
                rows=[{"a": "x"}],
                freeze_header=True,
            )
        ],
    )
    assert opened(render_spreadsheet(spec, options))["S"].freeze_panes == "A2"


def test_freezing_can_be_switched_off(options: RenderOptions) -> None:
    spec = SheetSpec(
        title="T",
        sheets=[
            Sheet(
                name="S",
                columns=[Column(header="A", key="a")],
                rows=[{"a": "x"}],
                freeze_header=False,
            )
        ],
    )
    assert opened(render_spreadsheet(spec, options))["S"].freeze_panes in (None, "A1")


def test_autofilter_covers_the_data_range(options: RenderOptions) -> None:
    spec = SheetSpec(
        title="T",
        sheets=[
            Sheet(
                name="S",
                columns=[Column(header="A", key="a")],
                rows=[{"a": "x"}, {"a": "y"}],
                autofilter=True,
            )
        ],
    )
    assert opened(render_spreadsheet(spec, options))["S"].auto_filter.ref is not None


def test_conditional_formatting_is_attached(options: RenderOptions) -> None:
    spec = SheetSpec(
        title="T",
        sheets=[
            Sheet(
                name="S",
                columns=[Column(header="Wert", key="w", type="number")],
                rows=[{"w": 1.0}, {"w": 2.0}],
                conditional_formats=[
                    ConditionalFormat(range="A2:A3", rule="color_scale")
                ],
            )
        ],
    )
    sheet = opened(render_spreadsheet(spec, options))["S"]
    assert len(list(sheet.conditional_formatting)) >= 1


def test_chart_is_anchored_to_the_sheet(options: RenderOptions) -> None:
    spec = SheetSpec(
        title="T",
        sheets=[
            Sheet(
                name="S",
                columns=[
                    Column(header="Monat", key="m"),
                    Column(header="Wert", key="w", type="number"),
                ],
                rows=[{"m": "Jan", "w": 1.0}, {"m": "Feb", "w": 2.0}],
                charts=[
                    SheetChart(
                        chart_type="line",
                        title="Verlauf",
                        categories_range="A2:A3",
                        values_range="B1:B3",
                        anchor="D2",
                    )
                ],
            )
        ],
    )
    assert len(opened(render_spreadsheet(spec, options))["S"]._charts) == 1


def test_column_widths_stay_within_bounds(options: RenderOptions) -> None:
    spec = SheetSpec(
        title="T",
        sheets=[
            Sheet(
                name="S",
                columns=[Column(header="A", key="a")],
                rows=[{"a": "x" * 500}],
            )
        ],
    )
    sheet = opened(render_spreadsheet(spec, options))["S"]
    width = sheet.column_dimensions["A"].width
    if width:
        assert 8.0 <= width <= 60.0
